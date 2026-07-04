"""CSV/XLSX 파일 파서 — 헤더 기반 행 dict 목록으로 변환한다."""

from __future__ import annotations

import csv
import io


class TabularParseError(Exception):
    """파일 파싱 실패."""


def parse_csv(content: bytes) -> list[dict[str, str]]:
    """CSV(UTF-8, BOM 허용) → 행 dict 목록."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp949")  # 사내 Excel 저장 CSV 대비
        except UnicodeDecodeError as exc:
            raise TabularParseError("CSV 인코딩 해석 실패 (UTF-8/CP949 아님)") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise TabularParseError("CSV 헤더가 없습니다")
    return [dict(row) for row in reader]


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    """XLSX 첫 번째 시트 → 행 dict 목록 (1행을 헤더로)."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise TabularParseError(f"XLSX 파싱 실패: {exc}") from exc
    sheet = workbook.active
    if sheet is None:
        raise TabularParseError("XLSX에 시트가 없습니다")
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration as exc:
        raise TabularParseError("XLSX 헤더가 없습니다") from exc
    result: list[dict[str, str]] = []
    for row in rows:
        if all(cell is None for cell in row):
            continue
        record = {
            header[index]: ("" if cell is None else str(cell))
            for index, cell in enumerate(row)
            if index < len(header) and header[index]
        }
        result.append(record)
    workbook.close()
    return result


def parse_tabular(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv(content)
    if lower.endswith(".xlsx"):
        return parse_xlsx(content)
    raise TabularParseError(f"지원하지 않는 파일 형식: {filename} (csv/xlsx만 지원)")
